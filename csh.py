import openpyxl
import csv
from argparse import (
    ArgumentParser,
)
import os
import re


# this utility when making questions uses specific format from
# the Kahoot template
# https://kahoot.com/blog/2018/08/23/import-kahoot-from-spreadsheet/
# the questions start at row 9 and in column B (2)
# Question, Answer 1, Answer 2, Answer 3, Answer 4, seconds, correct answers
# finally, I've added a final column for an explanation just in case
START_ROW = 9
START_COL = 2
NUM_COLS = 8

# ===================================
# helper functions for questions mode
# ===================================
def qdict_to_checkbox(question_dict: dict) -> str:
    # parse the correct answers into the correct format for the checkbox
    # the correct answers are in a comma separated
    correct = question_dict["solution"].split(",")
    # get rid of any extra spaces
    correct = [s.strip() for s in correct]
    # should turn an input like "1,2" into [True, True, False, False]
    # need the offset to account for the fact that the correct answers are
    # 1-indexed
    correct_TF = [str(x + 1) in correct for x in range(len(question_dict["options"]))]

    replacements = question_dict.copy()

    replacements["solution"] = correct_TF

    tmp = open("templates/multiple-select.catsoop", "r").read()
    return tmp.format(**replacements)


def qdict_to_radio(question_dict: dict) -> str:
    # parse the correct answers into the correct format for the
    replacements = question_dict.copy()

    # change from 1-indexing to 0-indexing
    replacements["solution"] = int(replacements["solution"]) - 1
    replacements["solution"] = replacements["options"][replacements["solution"]]

    tmp = open("templates/multiple-choice.catsoop", "r").read()
    return tmp.format(**replacements)


def qdict_to_shortanswer(question_dict: dict) -> str:
    # parse the correct answers into the correct format for the
    replacements = question_dict.copy()

    replacements["solution"] = [x.lower().strip() for x in replacements["options"]]

    tmp = open("templates/short-answer.catsoop", "r").read()
    return tmp.format(**replacements)


def questions_xlsx(wb: openpyxl.Workbook):
    ws = wb.active  # gets the worksheet with the data (sheet 1)

    for row in ws.iter_rows(
        min_row=START_ROW,
        min_col=START_COL,
        max_col=START_COL + NUM_COLS,
        values_only=True,
    ):

        # we've hit the end of the questions if the question is emkpty
        if not isinstance(row[0], str) or len(row[0].strip()) == 0:
            break

        # need to make sure everything is read in like a string
        exp = f"csq_explanation = '{row[7]}'" if row[7] else ""
        question_dict = {
            "question": str(row[0]),
            # if x to get rid of empty spaces
            "options": [str(x) for x in row[1:5] if x],
            "solution": str(row[6]) if row[6] else "",
            "explanation": exp,
        }

        # filter options that are empty
        question_dict["options"] = [x for x in question_dict["options"] if len(x)]

        # then this is a short answer question
        if len(question_dict["solution"].strip()) == 0:
            # print(question_dict["solution"], "gives a short answer question")
            question = qdict_to_shortanswer(question_dict)
            print(question)
        # then this is a multiple select question
        elif "," in question_dict["solution"]:
            # print(question_dict["solution"], "gives a multiple select answer question")
            question = qdict_to_checkbox(question_dict)
            print(question)
        # finally a radio question
        else:
            # print(question_dict["solution"], "gives a single select answer question")

            question = qdict_to_radio(question_dict)
            print(question)


def csv_to_xlsx(fname: str) -> openpyxl.Workbook:
    wb = openpyxl.Workbook()
    ws = wb.active

    with open("file.csv") as f:
        reader = csv.reader(f, delimiter=":")
        for row in reader:
            ws.append(row)

    return wb

# =================================================
# helper functions for number_text_blanks mode
# =================================================
def numberer(original_text: str) -> str:
    """
    Takes in the original_text string and outputs the string with the blanks numbered
    Places were the __([text])__ are considered blanks
    """
    # pattern from https://stackoverflow.com/questions/6208367/regex-to-match-stuff-between-parentheses but adjusted to look for strings with the '__'on either side as well
    new_text = original_text
    pattern = "__\(([^\)]+)\)__"

    replacements = []
    for i, m in enumerate(re.finditer(pattern, original_text)):
        b, f = m.start(), m.end()
        replacements.append(
            (b, f, original_text[b : b + 3] + str(i + 1) + ", " + original_text[b + 3 : f])
        )

    for b, f, new_string in reversed(replacements):
        new_text = new_text[:b] + new_string + new_text[f:]

    return new_text

# =================================================
# mode functions
# =================================================
def questions(fname: str):
    # if its an xslx, send it to the appropriate function
    # if its a CSV file, send it to the appropriate for decoding
    extension = fname.split(".")[-1]
    if extension == "xlsx":
        wb = openpyxl.load_workbook(filename=fname)
    elif extension == "csv":
        wb = csv_to_xlsx(fname)
    else:
        raise RuntimeError("invalid file type. takes csv or xlsx files")

    # the result is to print the questions to stdout
    questions_xlsx(wb)

def mkpg(dir_name: str, preload_name: str = ""):
    # make a new directoy
    os.mkdir(dir_name)
    # make a content.catsoop file
    open(dir_name + "/content.catsoop", "a").close()
    # make the preload.py file
    if not len(preload_name):
        preload_name = dir_name
    with open(dir_name + "/preload.py", "a") as f:
        f.write(f"cs_long_name = '{preload_name}'")
        f.close()

def number_text_blanks(fname: str):
    original_text = open(fname, 'r').read()
    new_text = numberer(original_text)
    print(new_text)


if __name__ == "__main__":
    # Parse command line arguments
    parser = ArgumentParser()

    # two mode options for parsing questions or for making a page
    parser.add_argument(
        "-q",
        "--question",
        default=None,
        help="""
        Create catsoop style questions from the input .xslx or .csv file.
        Format based on the excel template for making Kahoot questions
        """,
    )
    parser.add_argument(
        "-mkpg",
        default=None,
        nargs="+",
        help="""
        Create a new directory for a new catsoop page.
        First argument is taken as the directory name.
        Second (optional) argument is the name for the preload.py.
        """,
    )
    parser.add_argument(
        "-ntb",
        "--number-text-blanks",
        default=None,
        help="Numbers the blanks in a text with the format __([text])__. Give a readable text file.",
    )

    args, unknownargs = parser.parse_known_args()

    # verfiy the number of modes given as arguments
    notnone = lambda x: x is not None
    modes = (args.mkpg, args.question, args.number_text_blanks)
    if sum(map(notnone, modes)) > 1:
        raise RuntimeError("Too many modes given. ")
    if sum(map(notnone, modes)) == 0:
        raise RuntimeError("No mode given given. Please choose either -q or -mkpg")

    # decide what mode to jump into
    if args.question:
        questions(args.question)
    elif args.mkpg:
        if len(args.mkpg) > 1:
            mkpg(args.mkpg[0], args.mkpg[1])
        else:
            mkpg(args.mkpg[0])
    elif args.number_text_blanks:
        number_text_blanks(args.number_text_blanks)
